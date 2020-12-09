from openpyxl import load_workbook, Workbook
from django.core.management import BaseCommand
from django.db.utils import IntegrityError
from nomenclature.models import *


SERVICE_TYPES = [
    'Not defined',
    'ПРОФ',
    'Лабораторное исследование',
    'Коммерческий профиль',
    'Услуга'
]

class Command(BaseCommand):

    def handle(self, *args, **kwargs):

        ServiceType.objects.all().delete()

        for type in SERVICE_TYPES:
            new_type = ServiceType(name=type)
            new_type.save()

        Group.objects.all().delete()
        SubGroup.objects.all().delete()

        wb = load_workbook('nomenclature/data/Список групп и подгрупп.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]
        groups = {}

        for row in first_sheet.rows:
            if str(row[0].value) not in groups.keys():
                groups[str(row[0].value)] = {'name':str(row[1].value), 'subgroups':[]}
                groups[str(row[0].value)]['subgroups'].append({'number':str(row[2].value), 'name':str(row[3].value)})
                continue

            groups[str(row[0].value)]['subgroups'].append({'number':str(row[2].value), 'name':str(row[3].value)})

        for group in groups.keys():
            new_group = Group(number=group, name=groups[group]['name'])
            new_group.save()
            for sg in groups[group]['subgroups']:
                new_sg = SubGroup(number=sg['number'], name=sg['name'], group=new_group)
                new_sg.save()

        new_group = Group(number='99', name='не определена!')
        new_group.save()
        new_sg = SubGroup(number='99', name='не определена!', group=new_group)
        new_sg.save()

        Service.objects.all().delete()
        wb = load_workbook('nomenclature/data/nomenclature.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]

        for row in first_sheet:
            if row[0].value is not None:
                try:
                    gr = Group.objects.get(number=str(row[0].value))
                except:
                    print(f'группа {row[0].value} не найдена, для теста {row[7].value}')
            else:
                try:
                    gr = Group.objects.get(number='99')
                except:
                    print(f'группа 99 не найдена, для теста {row[7].value}')

            if row[1].value is not None:
                try:
                    sg = SubGroup.objects.get(number=str(row[1].value), group=gr)
                except:
                    print(f'подгруппа {row[1].value} не найдена, для теста {row[7].value}')
            else:
                try:
                    sg = SubGroup.objects.get(number='99')
                except:
                    print(f'подгруппа 99 не найдена, для теста {row[7].value}')

            if row[3].value is not None:
                type = ServiceType.objects.get(name=row[3].value)
            else:
                type = ServiceType.objects.get(name='Not defined')

            if type.name == 'Коммерческий профиль':
                new_record = Profile()
            else:
                new_record = Service()

            try:
                new_record.subgroup=sg
                new_record.salesability=True
                new_record.clients_group=row[2].value
                new_record.type=type
                new_record.classifier_1с=row[4].value
                new_record.tcle_code=row[5].value
                new_record.tcle_abbreviation=row[6].value
                new_record.code=row[7].value
                new_record.name=row[8].value
                new_record.blanks=row[9].value
                new_record.biomaterials=row[10].value
                new_record.container=row[11].value
                new_record.result_type=row[12].value
                new_record.due_date=row[13].value

                new_record.save()
            except IntegrityError:
                print(f'код {row[7].value} - не уникальный, второй раз не добавлен')


        profiles = Profile.objects.all()
        for profile in profiles:
            profile.services.clear()

        wb = load_workbook('nomenclature/data/profile.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]

        for row in first_sheet.rows:
            try:
                profile = Profile.objects.get(code=row[0].value)
            except:
                print(f'Профиля {row[0].value} нет в номенклатуре')

            try:
                service = Service.objects.get(code=row[1].value)
            except:
                print(f'Услуги {row[1].value} нет в номенклатуре')
                continue

            profile.services.add(service)

        wb = load_workbook('nomenclature/data/test_set.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]

        for row in first_sheet.rows:
            test_set = TestSet(key_code=row[1].value, name=row[2].value, department=row[3].value, addendum_key=row[4].value)
            test_set.save()

            try:
                service = Service.objects.get(code=row[0].value)
            except:
                print(f'Услуги {row[0].value} - нет в номенклатуре')

            service.test_set=test_set
            service.save()

        wb = load_workbook('nomenclature/data/test.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]

        for num, row in enumerate(first_sheet.rows):

            check_test = Test.objects.filter(keycode=row[0].value)
            if not check_test:
                test = Test(
                    keycode=row[0].value,
                    name=row[1].value,
                    short_name=row[1].value[:50],
                    result_type=row[4].value,
                    decimal_places=5,
                    kdl_test_key=row[2].value,
                    measure_unit=row[3].value,
                )

                test.save()

            if row[10].value is not None:
                test = Test.objects.get(keycode=row[0].value)
                new_reference = Reference(
                    test=test,
                    position=int(row[10].value[:-4])
                )

                if row[5].value is None:
                    new_reference.sex = 'Любой'
                if row[6].value is not None:
                    if '.' in row[6].value:
                        age = row[6].value.split('.')
                        yy = '00' if not age[0] else age[0]
                        mm = age[1][:2]
                        dd = age[1][2:]
                        age_from = f'{yy}:{mm}:{dd}'
                        new_reference.age_from = age_from
                    else:
                        new_reference.age_from = f'{row[6].value}:00:00'
                if row[7].value is not None:
                    if '.' in row[7].value:
                        age = row[7].value.split('.')
                        yy = '00' if not age[0] else age[0]
                        mm = age[1][:2]
                        dd = age[1][2:]
                        age_to = f'{yy}:{mm}:{dd}'
                        new_reference.age_to = age_to
                    else:
                        new_reference.age_to = f'{row[7].value}:00:00'
                if row[8].value is not None:
                    new_reference.lower_normal_value = row[8].value
                if row[9].value is not None:
                    new_reference.upper_normal_value = row[9].value
                if row[13].value is not None:
                    new_reference.normal_text = row[13].value
                if row[11].value is not None:
                    new_reference.normal_text = row[13].value
                if row[11].value is not None:
                    new_reference.clinic_interpretation_key = row[11].value
                if row[12].value is not None:
                    new_reference.clinic_interpretation_text = row[12].value

                new_reference.save()

        wb = load_workbook('nomenclature/data/med.xlsx', read_only=True)
        first_sheet = wb.worksheets[0]

        for row in first_sheet.rows:
            try:
                service = Service.objects.get(code=row[0].value)
            except:
                print(f'Услуги {row[0].value} - нет в номенклатуре')
                continue
            new_record = MadicineData(
                service=service,
                alter_name_KC=row[1].value,
                alter_name=row[2].value,
                note=row[3].value,
                volume_pp=row[4].value,
                container_pp=row[5].value,
                guide_pp=row[6].value,
                transport_conditions=row[7].value,
                term_assign=row[8].value,
                description=row[9].value,
                method=row[10].value,
                factors=row[11].value,
                preparation=row[12].value,
            )

            new_record.save()
